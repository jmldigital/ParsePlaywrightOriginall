import time
import re
import logging
import asyncio
from pathlib import Path
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By

from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from twocaptcha import TwoCaptcha
from dotenv import load_dotenv
import os

from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager

from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook

# Загрузка переменных окружения
load_dotenv()

# ==================== КОНФИГУРАЦИЯ ====================
API_KEY_2CAPTCHA = os.getenv('API_KEY_2CAPTCHA', 'your_api_key_here')
AVTO_LOGIN = os.getenv('AVTO_LOGIN', 'your_login_here')
AVTO_PASSWORD = os.getenv('AVTO_PASSWORD', 'your_password_here')

# Telegram настройки
BOT_TOKEN = os.getenv('BOT_TOKEN', '8364237483:AAERd9UAqQO_EAPt62AepFSojT41v9Vmw3s')
ADMIN_CHAT_ID = int(os.getenv('ADMIN_CHAT_ID', '-4688651319'))
SEND_TO_TELEGRAM = os.getenv('SEND_TO_TELEGRAM', 'True').lower() == 'true'

INPUT_FILE = 'input/наличие.xls'
OUTPUT_FILE = 'output/наличие_with_competitors.xlsx'
TEMP_FILE = 'output/наличие_temp.xlsx'
MAX_ROWS = 500
SAVE_INTERVAL = 10

# Таймауты и задержки
PAGE_LOAD_TIMEOUT = 60
DEFAULT_WAIT = 15
REQUEST_DELAY = 1.5
CAPTCHA_WAIT = 5

# Retry настройки
MAX_RETRIES = 3
RETRY_DELAY = 2

competitor1 = 'stparts_price'
competitor1_delivery = 'stparts_delivery'
competitor2 = 'avtoformula_price'
competitor2_delivery = 'avtoformula_delivery'
corrected_price='corrected_price'
input_price = 5

# CSS селекторы
SELECTORS = {
    'stparts': {
        'captcha_img': 'img.captchaImg',
        'captcha_input': "input[name='captcha']",
        'captcha_submit': 'captchaSubmitBtn',
        'results_table': 'table.globalResult.searchResultsSecondStep',
        'result_row': 'tr.resultTr2',
        'brand': 'td.resultBrand',
        'delivery': 'td.resultDeadline',
        'price': 'td.resultPrice'
    },
    'avtoformula': {
        'login_field': 'userlogin',
        'password_field': 'userpassword',
        'login_button': "input[type='submit'][name='login']",
        'article_field': 'article',
        'search_button': 'input[name="search"][data-action="ajaxSearch"]',
        'smode_select': 'smode',
        'results_table': 'table.web_ar_datagrid.search_results',
        'brand_cell': 'td.td_prd_info_link',
        'delivery_cell': 'td.td_term',
        'price_cell': 'td.td_final_price'
    }
}

# ==================== НАСТРОЙКА ЛОГИРОВАНИЯ ====================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('parser.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


# ==================== TELEGRAM ФУНКЦИИ ====================
async def send_telegram_file(file_path, caption=None):
    """Асинхронная отправка файла в Telegram"""
    from telegram import Bot
    
    try:
        bot = Bot(token=BOT_TOKEN)
        
        file_size = Path(file_path).stat().st_size / 1024  # KB
        
        default_caption = (
            f"✅ Обработка завершена!\n\n"
            f"📄 Файл: {Path(file_path).name}\n"
            f"📊 Размер: {file_size:.2f} KB"
        )
        
        with open(file_path, 'rb') as file:
            await bot.send_document(
                chat_id=ADMIN_CHAT_ID,
                document=file,
                caption=caption or default_caption,
                filename=Path(file_path).name
            )
        
        logger.info(f"✅ Файл {file_path} отправлен в Telegram (чат {ADMIN_CHAT_ID})")
        return True
        
    except Exception as e:
        logger.error(f"❌ Ошибка при отправке файла в Telegram: {e}")
        return False


def send_result_to_telegram(file_path, processed_count=0, total_count=0):
    """Синхронная обёртка для отправки файла"""
    if not SEND_TO_TELEGRAM:
        logger.info("Отправка в Telegram отключена (SEND_TO_TELEGRAM=False)")
        return
    
    caption = (
        f"✅ Обработка завершена!\n\n"
        f"📄 Файл: {Path(file_path).name}\n"
        f"📊 Обработано: {processed_count}/{total_count} позиций\n"
        f"📦 Размер: {Path(file_path).stat().st_size / 1024:.2f} KB"
    )
    
    try:
        asyncio.run(send_telegram_file(file_path, caption))
    except Exception as e:
        logger.error(f"Ошибка при отправке в Telegram: {e}")


# ==================== ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ====================
def setup_driver():
    """Настройка и создание WebDriver"""
    options = webdriver.ChromeOptions()
    options.add_argument("--headless=False")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option('useAutomationExtension', False)
    options.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/115.0.0.0 Safari/537.36"
    )
    
    try:
        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=options)
        driver.set_page_load_timeout(PAGE_LOAD_TIMEOUT)
        logger.info("WebDriver успешно инициализирован")
        return driver
    except Exception as e:
        logger.error(f"Ошибка при создании WebDriver: {e}")
        raise


def parse_price(text):
    """Парсинг цены из текста"""
    if not text or not isinstance(text, str):
        return None
    
    clean = re.sub(r"[^\d,\.\s]", "", text).strip()
    clean = clean.replace(" ", "")
    
    if clean.count(",") == 1 and clean.count(".") == 0:
        clean = clean.replace(",", ".")
    elif clean.count(",") > 1:
        clean = clean.replace(",", "")
    
    try:
        return float(clean)
    except (ValueError, AttributeError):
        logger.warning(f"Не удалось распарсить цену: {text}")
        return None


def preprocess_dataframe(df):
    """Удаляет слеши из колонки бренда"""
    try:
        brand_col_idx = 2
        if len(df.columns) > brand_col_idx:
            df.iloc[:, brand_col_idx] = (
                df.iloc[:, brand_col_idx]
                  .astype(str)
                  .str.replace('/', '', regex=False)
                  .str.replace('\\', '', regex=False)
                  .str.strip()
            )
    except Exception as e:
        logger.error(f"Ошибка при предобработке данных: {e}")
    return df


def normalize_brand(brand_str):
    """Нормализация названия бренда для сравнения"""
    if not brand_str:
        return ""
    return re.sub(r'[^a-z0-9]', '', str(brand_str).lower())


def brand_matches(search_brand, result_brand):
    """Проверка соответствия бренда (двусторонняя)"""
    if not search_brand or not result_brand:
        return False
    
    norm_search = normalize_brand(search_brand)
    norm_result = normalize_brand(result_brand)
    
    return norm_search in norm_result or norm_result in norm_search


def save_progress(df, filename=TEMP_FILE):
    """Сохранение промежуточного прогресса"""
    try:
        df.to_excel(filename, index=False)
        logger.info(f"Промежуточный прогресс сохранён в {filename}")
    except Exception as e:
        logger.error(f"Ошибка при сохранении прогресса: {e}")


def retry_on_failure(func, *args, max_retries=MAX_RETRIES, **kwargs):
    """Декоратор для повторных попыток при ошибке"""
    for attempt in range(max_retries):
        try:
            return func(*args, **kwargs)
        except Exception as e:
            if attempt < max_retries - 1:
                logger.warning(f"Попытка {attempt + 1}/{max_retries} не удалась: {e}. Повтор через {RETRY_DELAY}с")
                time.sleep(RETRY_DELAY)
            else:
                logger.error(f"Все {max_retries} попытки исчерпаны для {func.__name__}")
                raise


# ==================== РАБОТА С КАПЧЕЙ ====================
def solve_image_captcha(driver):
    """Решение капчи через 2Captcha"""
    try:
        solver = TwoCaptcha(API_KEY_2CAPTCHA)
        
        img_el = driver.find_element(By.CSS_SELECTOR, SELECTORS['stparts']['captcha_img'])
        captcha_base64 = img_el.screenshot_as_base64

        logger.info("Отправляем капчу на распознавание в 2Captcha")
        result = solver.normal(captcha_base64)
        captcha_text = result['code']
        logger.info(f"Капча распознана: {captcha_text}")

        input_el = driver.find_element(By.CSS_SELECTOR, SELECTORS['stparts']['captcha_input'])
        input_el.clear()
        input_el.send_keys(captcha_text)

        submit_btn = driver.find_element(By.ID, SELECTORS['stparts']['captcha_submit'])
        submit_btn.click()
        time.sleep(CAPTCHA_WAIT)

        return True
    except Exception as e:
        logger.error(f"Ошибка решения капчи: {e}")
        return False


# ==================== ПАРСИНГ STPARTS.RU ====================
def scrape_stparts(driver, brand, part):
    """Парсинг данных с stparts.ru"""
    url = f"https://stparts.ru/search/{brand}/{part}"
    
    try:
        driver.get(url)
        logger.info(f"Загружена страница: {url}")
        
        if len(driver.find_elements(By.CSS_SELECTOR, SELECTORS['stparts']['captcha_img'])) > 0:
            logger.warning("Обнаружена капча на stparts.ru")
            if not solve_image_captcha(driver):
                raise Exception("Не удалось решить капчу")

        wait = WebDriverWait(driver, DEFAULT_WAIT)

        wait.until(EC.presence_of_element_located(
            (By.CSS_SELECTOR, SELECTORS['stparts']['results_table'])))
        
        table = driver.find_element(By.CSS_SELECTOR, SELECTORS['stparts']['results_table'])
        wait.until(lambda d: len(d.find_elements(
            By.CSS_SELECTOR, SELECTORS['stparts']['result_row'])) > 0)
        
        rows = table.find_elements(By.CSS_SELECTOR, SELECTORS['stparts']['result_row'])
        
        if not rows:
            logger.info(f"Результаты не найдены для {brand} / {part}")
            return None, None

        logger.info(f"Найдено {len(rows)} строк результатов")

        for priority_search in [True, False]:
            for row in rows:
                try:
                    brand_td = row.find_element(By.CSS_SELECTOR, SELECTORS['stparts']['brand'])
                    brand_in_row = brand_td.text.strip()
                    
                    if not brand_matches(brand, brand_in_row):
                        continue

                    delivery_td = row.find_element(By.CSS_SELECTOR, SELECTORS['stparts']['delivery'])
                    delivery_min = delivery_td.text.strip()

                    if priority_search and not re.match(r"^1(\D|$)", delivery_min):
                        continue

                    price_text = row.find_element(By.CSS_SELECTOR, SELECTORS['stparts']['price']).text
                    price = parse_price(price_text)
                    
                    if price is not None:
                        if priority_search:
                            logger.info(f"Найдено в наличии (бренд: {brand_in_row}, срок {delivery_min}): {price} ₽")
                        else:
                            logger.info(f"Найдено (бренд: {brand_in_row}, срок {delivery_min}): {price} ₽")
                        return price, delivery_min
                        
                except NoSuchElementException:
                    continue

        logger.info(f"Подходящие результаты не найдены для {brand} / {part}")
        return None, None

    except TimeoutException:
        logger.error(f"Timeout при загрузке результатов для {brand} / {part}")
        return None, None
    except Exception as e:
        logger.error(f"Ошибка парсинга stparts для {brand} / {part}: {e}")
        return None, None


# ==================== АВТОРИЗАЦИЯ AVTOFORMULA ====================
def login_avtoformula(driver, login, password):
    """Авторизация на avtoformula.ru"""
    login_url = "https://www.avtoformula.ru"
    
    try:
        driver.get(login_url)
        wait = WebDriverWait(driver, DEFAULT_WAIT)

        logger.info("Начинаем авторизацию на avtoformula.ru")
        
        login_el = wait.until(EC.element_to_be_clickable(
            (By.ID, SELECTORS['avtoformula']['login_field'])))
        login_el.clear()
        login_el.send_keys(login)

        password_el = driver.find_element(By.ID, SELECTORS['avtoformula']['password_field'])
        password_el.clear()
        password_el.send_keys(password)

        submit_btn = driver.find_element(By.CSS_SELECTOR, SELECTORS['avtoformula']['login_button'])
        submit_btn.click()

        wait.until(EC.invisibility_of_element((By.ID, SELECTORS['avtoformula']['login_field'])))
        logger.info("Авторизация успешна")

        time.sleep(2)
        wait.until(EC.element_to_be_clickable((By.ID, SELECTORS['avtoformula']['smode_select'])))
        smode_select = driver.find_element(By.ID, SELECTORS['avtoformula']['smode_select'])
        
        for option in smode_select.find_elements(By.TAG_NAME, "option"):
            if option.get_attribute("value") == "A0":
                option.click()
                logger.info("Опция 'без аналогов' установлена")
                break

        return True
        
    except TimeoutException:
        logger.error("Timeout при авторизации на avtoformula.ru")
        return False
    except Exception as e:
        logger.error(f"Ошибка авторизации на avtoformula.ru: {e}")
        return False


# ==================== ПАРСИНГ AVTOFORMULA.RU ====================
def scrape_avtoformula(driver, brand, part):
    """Парсинг данных с avtoformula.ru"""
    try:
        wait = WebDriverWait(driver, DEFAULT_WAIT)
        driver.get("https://www.avtoformula.ru")

        article_field = wait.until(EC.element_to_be_clickable(
            (By.ID, SELECTORS['avtoformula']['article_field'])))
        article_field.clear()
        article_field.send_keys(part)
        logger.info(f"Введён артикул: {part}")

        search_btn = driver.find_element(By.CSS_SELECTOR, SELECTORS['avtoformula']['search_button'])
        search_btn.click()

        wait.until(EC.presence_of_element_located(
            (By.CSS_SELECTOR, SELECTORS['avtoformula']['results_table'])))
        
        rows = driver.find_elements(By.CSS_SELECTOR, 
            f"{SELECTORS['avtoformula']['results_table']} tr")
        
        if not rows or len(rows) < 2:
            logger.info(f"Результаты не найдены для {brand} / {part}")
            return None, None

        min_delivery = None
        min_price = None

        for row in rows[1:]:
            try:
                brand_td = row.find_element(By.CSS_SELECTOR, SELECTORS['avtoformula']['brand_cell'])
                brand_in_row = brand_td.text.strip()
                
                if not brand_matches(brand, brand_in_row):
                    continue

                delivery_td = row.find_element(By.CSS_SELECTOR, SELECTORS['avtoformula']['delivery_cell'])
                delivery_text = delivery_td.text.strip().split('/')[0].strip()
                
                try:
                    delivery_days = int(delivery_text)
                except (ValueError, IndexError):
                    continue

                price_td = row.find_element(By.CSS_SELECTOR, SELECTORS['avtoformula']['price_cell'])
                price = parse_price(price_td.text.strip())
                
                if price is None:
                    continue

                if min_delivery is None or delivery_days < min_delivery:
                    min_delivery = delivery_days
                    min_price = price
                    
            except NoSuchElementException:
                continue

        if min_delivery is not None:
            logger.info(f"Найдено: {min_delivery} дней, цена {min_price} ₽")
            return min_price, f"{min_delivery} дней"
        
        logger.info(f"Подходящие результаты не найдены для {brand} / {part}")
        return None, None

    except TimeoutException:
        logger.error(f"Timeout при поиске на avtoformula для {brand} / {part}")
        return None, None
    except Exception as e:
        logger.error(f"Ошибка парсинга avtoformula для {brand} / {part}: {e}")
        return None, None




def adjust_prices_and_save(df, output_file, original_price_col=input_price, competitor_price_cols=(competitor1, competitor2), competitor_delivery_cols=(competitor1_delivery, competitor2_delivery), corrected_col=corrected_price):
    """
    Добавляет столбец с корректировочной ценой и сохраняет DataFrame с цветовым форматированием.
    
    :param df: DataFrame с данными
    :param output_file: Путь для сохранения xlsx
    :param original_price_col: Название колонки с вашей ценой
    :param competitor_price_cols: Кортеж с названиями колонок конкурентов (цен)
    :param competitor_delivery_cols: Кортеж с названиями колонок конкурентов (доставок)
    :param corrected_col: Имя нового столбца для скорректированной цены
    """

    def parse_delivery_days(delivery_str):
        """Парсит срок доставки 'X дней' в int или None"""
        if not delivery_str or not isinstance(delivery_str, str):
            return None
        m = re.search(r'(\d+)', delivery_str)
        if m:
            return int(m.group(1))
        return None

    corrected_prices = []
    for idx, row in df.iterrows():
        try:
            our_price = float(row[original_price_col]) if pd.notna(row[original_price_col]) else None
            # print('our_price-----------',our_price)
            competitor_prices = []
            for price_col, delivery_col in zip(competitor_price_cols, competitor_delivery_cols):
                comp_price = row.get(price_col)

                # print('comp_price-----------',comp_price)
                # print(df.columns.tolist())

                comp_delivery = row.get(delivery_col)
                if pd.isna(comp_price) or pd.isna(comp_delivery):
                    continue
                comp_price = float(comp_price)
                delivery_days = parse_delivery_days(comp_delivery)
                if delivery_days is None:
                    continue
                # Условие: срок 1-4 дня — считаем цену
                if 1 <= delivery_days <= 4:
                    competitor_prices.append(comp_price)
            if not competitor_prices or our_price is None:
                # Нет данных для корректировки
                corrected_prices.append(our_price)
                continue
            
            min_competitor_price = min(competitor_prices)
            
            # Если минимальная конкурентная цена меньше нашей - уменьшить на 2 рубля
            if our_price > min_competitor_price:
                new_price = round(min_competitor_price - 2, 2)
                if new_price < 0:
                    new_price = 0.0
                corrected_prices.append(new_price)
            else:
                corrected_prices.append(our_price)
                
        except Exception as e:
            logger.error(f"Ошибка при вычислении скорректированной цены для строки {idx}: {e}")
            corrected_prices.append(row[original_price_col])
    
    df[corrected_col] = corrected_prices

    # пересохраняем DataFrame чтобы колонка и данные появилась в файле
    df.to_excel(output_file, index=False)

    # print(df)

    # Сохраняем с форматированием
    wb = load_workbook(output_file)
    ws = wb.active
    
    # Найдём индекс колонки с corrected_col
    max_col = ws.max_column
    max_row = ws.max_row
    header_row = 1
    col_idx = None
    for col in range(1, max_col + 1):
        cell_val = ws.cell(row=header_row, column=col).value
        if cell_val == corrected_col:
            col_idx = col
            break
    if col_idx is None:
        logger.warning(f"Колонка {corrected_col} не найдена при форматировании")
        wb.save(output_file)
        return

    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')  # красный
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # зелёный

    # Цвет для ячеек: красный, если цена ниже оригинала; зелёный — если не меняется
    for row in range(2, max_row + 1):
        original_price = ws.cell(row=row, column=ws.min_column + df.columns.get_loc(original_price_col)).value
        corrected_price = ws.cell(row=row, column=col_idx).value
        if original_price is None or corrected_price is None:
            continue
        try:
            if float(corrected_price) < float(original_price):
                ws.cell(row=row, column=col_idx).fill = red_fill
            else:
                ws.cell(row=row, column=col_idx).fill = green_fill
        except:
            pass

    wb.save(output_file)
    logger.info(f"Файл сохранён с корректировкой цен и форматированием: {output_file}")





# ==================== ОСНОВНАЯ ЛОГИКА ====================
def main():
    """Главная функция"""
    logger.info("=" * 60)
    logger.info("ЗАПУСК ПАРСЕРА АВТОЗАПЧАСТЕЙ")
    logger.info("=" * 60)
    
    # Проверка наличия входного файла
    if not Path(INPUT_FILE).exists():
        logger.error(f"Файл {INPUT_FILE} не найден!")
        return
    
    # Проверка credentials
    if API_KEY_2CAPTCHA == 'your_api_key_here':
        logger.warning("⚠️ API ключ 2Captcha не настроен!")
    if AVTO_LOGIN == 'your_login_here':
        logger.warning("⚠️ Логин Avtoformula не настроен!")
    
    # Загрузка данных
    try:
        df = pd.read_excel(INPUT_FILE)
        logger.info(f"Загружено {len(df)} строк из {INPUT_FILE}")

        df = preprocess_dataframe(df)

    except Exception as e:
        logger.error(f"Ошибка при чтении файла: {e}")
        return
    


    # Добавление новых колонок
    for col in [competitor1, competitor1_delivery, competitor2, competitor2_delivery]:
        if col not in df.columns:
            df[col] = ""
    
    driver = None
    processed_count = 0
    
    try:
        driver = setup_driver()
        
        if not login_avtoformula(driver, AVTO_LOGIN, AVTO_PASSWORD):
            logger.error("Не удалось авторизоваться на avtoformula.ru")
            return
        
        rows_to_process = df.head(MAX_ROWS)
        total_rows = len(rows_to_process)
        
        for idx, row in rows_to_process.iterrows():
            try:
                part = str(row[1]).strip()
                brand = str(row[3]).strip()
                
                if not part or not brand or part == 'nan' or brand == 'nan':
                    logger.warning(f"Строка {idx + 1}: пропущена (пустые данные)")
                    continue
                
                logger.info(f"\n{'=' * 60}")
                logger.info(f"Обработка {processed_count + 1}/{total_rows}: {brand} / {part}")
                logger.info(f"{'=' * 60}")
                
                logger.info("→ Поиск на stparts.ru")
                price_st, delivery_st = retry_on_failure(scrape_stparts, driver, brand, part)
                
                if price_st is not None:
                    df.at[idx, competitor1] = round(price_st, 2)
                if delivery_st is not None:
                    df.at[idx, competitor1_delivery] = delivery_st
                
                logger.info("→ Поиск на avtoformula.ru")
                price_avto, delivery_avto = retry_on_failure(scrape_avtoformula, driver, brand, part)
                
                if price_avto is not None:
                    df.at[idx, competitor2] = round(price_avto, 2)
                if delivery_avto is not None:
                    df.at[idx, competitor2_delivery] = delivery_avto
                
                processed_count += 1
                
                if processed_count % SAVE_INTERVAL == 0:
                    save_progress(df)
                
                time.sleep(REQUEST_DELAY)
                
            except Exception as e:
                logger.error(f"Критическая ошибка при обработке строки {idx + 1}: {e}")
                save_progress(df)
                continue
        
        logger.info(f"\n{'=' * 60}")
        logger.info(f"Обработка завершена: {processed_count}/{total_rows} строк")
        logger.info(f"{'=' * 60}")
        
    except KeyboardInterrupt:
        logger.warning("\n⚠️ Прервано пользователем")
        save_progress(df)
        
    except Exception as e:
        logger.error(f"Критическая ошибка: {e}", exc_info=True)
        save_progress(df)
        
    finally:
        if driver:
            driver.quit()
            logger.info("WebDriver закрыт")
        
        # Финальное сохранение
        try:
            df.to_excel(OUTPUT_FILE, index=False)
            logger.info(f"✅ Результаты сохранены в {OUTPUT_FILE}")
            

            adjust_prices_and_save(df, OUTPUT_FILE, original_price_col=input_price, competitor_price_cols=(competitor1, competitor2), competitor_delivery_cols=(competitor1_delivery, competitor2_delivery), corrected_col=corrected_price)

            # Отправка результата в Telegram
            if Path(OUTPUT_FILE).exists():
                logger.info("📤 Отправка результата в Telegram...")
                send_result_to_telegram(OUTPUT_FILE, processed_count, total_rows)
            
        except Exception as e:
            logger.error(f"Ошибка при сохранении финального файла: {e}")


if __name__ == "__main__":
    main()