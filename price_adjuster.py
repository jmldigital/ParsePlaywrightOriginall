# price_adjuster.py
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
import pandas as pd
from config import (
    corrected_price,
    stparts_price,
    avtoformula_price,
    stparts_delivery,
    avtoformula_delivery,
    input_price,
    INPUT_COL_ARTICLE,
    INPUT_COL_BRAND,
)

from utils import logger, parse_price
import re


def parse_delivery_days(delivery_str):
    """Парсит срок доставки 'X дней' в int или None"""
    if not delivery_str or not isinstance(delivery_str, str):
        return None
    m = re.search(r"(\d+)", delivery_str)
    if m:
        return int(m.group(1))
    return None


def adjust_prices_and_save(df, output_file):
    """
    Добавляет скорректированную цену и сохраняет DataFrame с цветовым форматированием.
    Подробно логирует процесс обработки строк и вычислений.
    """

    corrected_prices = []

    # logger.info(f"Типы датафрейма:\n{df}")

    for idx, row in df.iterrows():

        # logger.info(f"▶️ Обработка строки {idx + 1}/{len(df)}")

        try:
            # Наша цена
            raw_price = row.get(input_price)

            # Диагностика - что именно в переменной
            # logger.info(f"raw_price = {raw_price}, тип: {type(raw_price)}, repr: {repr(raw_price)}")

            our_price = parse_price(raw_price)
            # logger.info(f"our_price = {our_price}")

            if our_price is None:
                logger.warning(f"⚠️ Невозможно преобразовать цену: {raw_price}")
                corrected_prices.append(None)
                continue

            # logger.info(f"✅ our_price (float): {our_price}")

            # Цены конкурентов (только с доставкой 1–4 дня)
            competitor_prices = []

            # Проверяем competitor1
            comp1_price = parse_price(row.get(stparts_price))
            comp1_delivery = row.get(stparts_delivery)
            # logger.info(f"competitor1_price: {comp1_price}, delivery: {comp1_delivery}")

            if comp1_price is not None and pd.notna(comp1_delivery):
                delivery_days = parse_delivery_days(comp1_delivery)
                if delivery_days is not None and 1 <= delivery_days <= 4:
                    competitor_prices.append(comp1_price)
                    # logger.info(f"✅ competitor1 добавлен ({comp1_price} за {delivery_days} дн.)")

            # Проверяем competitor2
            comp2_price = parse_price(row.get(avtoformula_price))
            comp2_delivery = row.get(avtoformula_delivery)
            # logger.info(f"competitor2_price: {comp2_price}, delivery: {comp2_delivery}")

            if comp2_price is not None and pd.notna(comp2_delivery):
                delivery_days = parse_delivery_days(comp2_delivery)
                if delivery_days is not None and 1 <= delivery_days <= 4:
                    competitor_prices.append(comp2_price)
                    # logger.info(f"✅ competitor2 добавлен ({comp2_price} за {delivery_days} дн.)")

            # Логика корректировки
            if not competitor_prices:
                corrected_prices.append(our_price)
                # logger.info("ℹ️ Нет подходящих конкурентов, цена без изменений.")
            else:
                min_comp_price = min(competitor_prices)
                if our_price > min_comp_price:
                    new_price = max(int(min_comp_price - 2), int(our_price * 0.1), 1)
                    corrected_prices.append(new_price)
                else:
                    corrected_prices.append(our_price)
                    # logger.info(f"✅ Цена конкурентоспособна: {our_price}")

        except Exception as e:
            logger.error(
                f"Ошибка при вычислении скорректированной цены для строки {idx}: {e}"
            )
            corrected_prices.append(None)

    # Добавляем колонку
    df[corrected_price] = corrected_prices

    # logger.info(f"датафрейм перед всеми манипуляуиями {df[corrected_price]})")

    # Сохраняем с форматированием
    try:
        df.to_excel(output_file, index=False, engine="openpyxl")
        wb = load_workbook(output_file)
        ws = wb.active

        corr_col_idx = None
        orig_col_idx = None

        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if header == corrected_price:
                corr_col_idx = col
            if header == input_price:
                orig_col_idx = col

        # logger.info(f"Строка с индеком перед проверками {corr_col_idx}:значение с корректировочной ценой {corr_val_raw})")

        if corr_col_idx is None or orig_col_idx is None:
            logger.warning("⚠️ Не удалось определить индексы колонок для подсветки")
            wb.save(output_file)
            return

        red_fill = PatternFill(
            start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
        )  # снижена
        green_fill = PatternFill(
            start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
        )  # без изменений

        for row_idx in range(2, ws.max_row + 1):
            try:
                orig_val_raw = ws.cell(row=row_idx, column=orig_col_idx).value
                corr_val_raw = ws.cell(row=row_idx, column=corr_col_idx).value

                logger.debug(
                    f"Строка с индеком {corr_col_idx}:значение с корректировочной ценой {corr_val_raw})"
                )

                if orig_val_raw is None or corr_val_raw is None:
                    logger.debug(
                        f"Строка {row_idx}: пропущена из-за None значения (orig: {orig_val_raw}, corr: {corr_val_raw})"
                    )
                    continue

                orig_val = parse_price(orig_val_raw)
                corr_val = corr_val_raw

                logger.debug(
                    f"Строка {row_idx}: после parse_price -> orig_val: {orig_val}, corr_val: {corr_val}"
                )

                if orig_val is None or corr_val is None:
                    logger.debug(
                        f"Строка {row_idx}: пропущена из-за None после parse_price"
                    )
                    continue

                if corr_val < orig_val:
                    ws.cell(row=row_idx, column=corr_col_idx).fill = red_fill
                    logger.debug(
                        f"Строка {row_idx}: corr_val < orig_val, применён красный цвет"
                    )
                else:
                    ws.cell(row=row_idx, column=corr_col_idx).fill = green_fill
                    logger.debug(
                        f"Строка {row_idx}: corr_val >= orig_val, применён зелёный цвет"
                    )

            except Exception as cell_e:
                logger.error(f"Ошибка форматирования строки {row_idx}: {cell_e}")
                continue

        wb.save(output_file)
        logger.info(
            f"✅ Файл сохранён с корректировкой цен и форматированием: {output_file}"
        )

    except Exception as e:
        logger.error(f"❌ Ошибка при сохранении Excel с форматированием: {e}")
